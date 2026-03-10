import io
from django.http import HttpResponse
from django.utils import timezone
from apps.accounts.models import UserProfile
from apps.documents.models import Documento, TipoDocumento


def exportar_excel(clientes_data, total_tipos):
    """Genera reporte Excel de cumplimiento por cliente."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cumplimiento'

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Reporte de Cumplimiento — Acreditaciones Mineras'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10)

    # Encabezados
    headers = ['Cliente', 'Empresa', 'RUT', 'Docs Aprobados', 'Total Requeridos', 'Cumplimiento %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row, item in enumerate(clientes_data, 5):
        perfil = item['perfil']
        ws.cell(row=row, column=1, value=perfil.user.get_full_name() or perfil.user.username)
        ws.cell(row=row, column=2, value=perfil.empresa or '—')
        ws.cell(row=row, column=3, value=perfil.rut or '—')
        ws.cell(row=row, column=4, value=item['aprobados'])
        ws.cell(row=row, column=5, value=total_tipos)
        pct_cell = ws.cell(row=row, column=6, value=f"{item['cumplimiento']}%")
        pct_cell.alignment = Alignment(horizontal='center')

        # Color según cumplimiento
        if item['cumplimiento'] == 100:
            color = 'c8e6c9'
        elif item['cumplimiento'] >= 50:
            color = 'fff9c4'
        else:
            color = 'ffcdd2'
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=color, end_color=color, fill_type='solid'
            )

    # Ancho de columnas
    anchos = [30, 25, 15, 18, 18, 16]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Respuesta HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fecha = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cumplimiento_{fecha}.xlsx"'
    return response


def exportar_pdf(clientes_data, total_tipos):
    """Genera reporte PDF de cumplimiento por cliente."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    elementos = []

    # Título
    titulo_style = ParagraphStyle(
        'titulo',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=5
    )
    elementos.append(Paragraph('Reporte de Cumplimiento — Acreditaciones Mineras', titulo_style))
    elementos.append(Paragraph(
        f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}',
        styles['Normal']
    ))
    elementos.append(Spacer(1, 0.5*cm))

    # Tabla
    data = [['Cliente', 'Empresa', 'RUT', 'Aprobados', 'Requeridos', 'Cumplimiento']]

    for item in clientes_data:
        perfil = item['perfil']
        data.append([
            perfil.user.get_full_name() or perfil.user.username,
            perfil.empresa or '—',
            perfil.rut or '—',
            str(item['aprobados']),
            str(total_tipos),
            f"{item['cumplimiento']}%",
        ])

    tabla = Table(data, colWidths=[6*cm, 5*cm, 3.5*cm, 3*cm, 3*cm, 3*cm])

    estilo = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Datos
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (3, 1), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])

    # Color por cumplimiento
    for i, item in enumerate(clientes_data, 1):
        if item['cumplimiento'] == 100:
            color = colors.HexColor('#c8e6c9')
        elif item['cumplimiento'] >= 50:
            color = colors.HexColor('#fff9c4')
        else:
            color = colors.HexColor('#ffcdd2')
        estilo.add('BACKGROUND', (5, i), (5, i), color)

    tabla.setStyle(estilo)
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    fecha = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cumplimiento_{fecha}.pdf"'
    return response